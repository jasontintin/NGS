import glob
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

def count_variants(folder_path, output_file="variant_summary_Deletion.xlsx"):
    file_list = glob.glob(os.path.join(folder_path, "**/variants_Tit*.tsv"), recursive=True)

    records = []
    total_mutations = 0
    total_deletions = 0
    total_high_af_deletions = 0
    total_supFdeletion_96 = 0
    total_supFdeletion_97 = 0
    for file in file_list:
        try:
            df = pd.read_csv(file, sep='\t', encoding='utf-8')
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(file, sep='\t', encoding='ISO-8859-1')
            except Exception as e:
                print(f"Error reading {file}: {e}")
                continue

        if df.empty or'4. Start - start position of the variation' not in df.columns or '34. VARTYPE - variant type' not in df.columns or '15. AF - allele frequency' not in df.columns:
            print(f"Skipping invalid file (missing column): {file}")
            continue

        sample_name = os.path.basename(file).replace("variants_", "").replace(".tsv", "")
        num_variants = len(df)
        num_deletion = (df['34. VARTYPE - variant type'].str.lower() == 'deletion').sum()

        # 新增：篩選符合條件的行
        filtered_df = df[
            (df['34. VARTYPE - variant type'].str.lower() == 'deletion') &
            (df['15. AF - allele frequency'] > 0.6)
        ]
        high_af_deletions_counts  = len(filtered_df)

        num_supFDeletion_96_df = df[
            (df['34. VARTYPE - variant type'].str.lower() == 'deletion') &
            (df['4. Start - start position of the variation'] == 191)
        ]
        supFDeletion_96_counts  = len(num_supFDeletion_96_df)

        num_supFDeletion_97_df = df[
            (df['34. VARTYPE - variant type'].str.lower() == 'deletion') &
            (df['4. Start - start position of the variation'] == 192)
            ]
        supFDeletion_97_counts = len(num_supFDeletion_97_df)

        total_mutations += num_variants
        total_deletions += num_deletion
        total_high_af_deletions += high_af_deletions_counts
        total_supFdeletion_96 += supFDeletion_96_counts
        total_supFdeletion_97 += supFDeletion_97_counts

        print(
            f'Processed: {sample_name} - Total: {num_variants}, Deletions: {num_deletion}, High AF Deletion Samples (>0.6): {high_af_deletions_counts},/n'
            f'supFDeletion_96_counts: {supFDeletion_96_counts}, supFDeletion_97_counts: {supFDeletion_97_counts}')

        records.append({
            "Sample": sample_name,
            "Total Variants": num_variants,
            "Deletions": num_deletion,
            "High AF Deletions (>0.6)": high_af_deletions_counts,
            "supFDeletion_96_counts": supFDeletion_96_counts,
            "supFDeletion_97_counts": supFDeletion_97_counts,
        })

    result_df = pd.DataFrame(records)
    result_df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    title_font = Font(name="Arial", size=14, bold=True)
    for cell in ws[1]:
        cell.font = title_font

    data_font = Font(name="Arial", size=12)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font

    wb.save(output_file)
    print(f"\nResults saved to {output_file} with formatted font.")

# 執行
count_variants('/Users/jasontingting/Desktop/K3 部分結果', '/Users/jasontingting/Desktop/variant_summary_Deletion数.xlsx')