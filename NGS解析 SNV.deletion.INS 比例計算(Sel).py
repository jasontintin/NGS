import glob
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def count_variants(folder_path, output_file="variant_summary_supF.xlsx"):
    file_list = glob.glob(os.path.join(folder_path, "**/variants_*.tsv"), recursive=True)

    records = []
    for file in file_list:
        try:
            df = pd.read_csv(file, sep='\t', encoding='utf-8')
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(file, sep='\t', encoding='ISO-8859-1')
            except Exception as e:
                print(f"Error reading {file}: {e}")
                continue

        required_cols = [
            '4. Start - start position of the variation',
            '34. VARTYPE - variant type',
            '15. AF - allele frequency',
            '1. Sample - sample name'
        ]
        if not all(col in df.columns for col in required_cols):
            print(f"Skipping invalid file (missing column): {file}")
            continue

        sample_name = os.path.basename(file).replace("variants_", "").replace(".tsv", "")
        df['Start'] = df['4. Start - start position of the variation']
        df['AF'] = df['15. AF - allele frequency']
        df['VARTYPE'] = df['34. VARTYPE - variant type'].str.lower()
        df['SampleName'] = df['1. Sample - sample name']

        # 只取 start > 96 的 supF 區域
        df_supf = df[df['Start'] > 99]

        # 額外條件：排除 LNATC 文件中 Start=192 且 VARTYPE=snv 的變異
        if 'LNATC' in sample_name:
            df_supf = df_supf[~((df_supf['Start'] == 192) & (df_supf['VARTYPE'] == 'snv'))]

        af_thresholds = [0.01, 0.4, 0.5, 0.6]
        variant_types = ['deletion', 'insertion', 'snv']

        record = {"Sample": sample_name}

        for af in af_thresholds:
            af_key = f"AF>={af}"

            # 全部 supF 變異數
            total_count = len(df_supf[df_supf['AF'] >= af])
            record[f"Total_{af_key}"] = total_count

            # 分類統計：deletion / insertion / snv
            for vtype in variant_types:
                count = len(df_supf[
                    (df_supf['AF'] >= af) &
                    (df_supf['VARTYPE'] == vtype)
                ])
                vtype_key = vtype.upper()
                record[f"{vtype_key}_{af_key}"] = count

                # 百分比 (避免除以 0)
                if total_count > 0:
                    ratio = count / total_count
                else:
                    ratio = 0
                record[f"{vtype_key}_ratio_{af_key}"] = ratio

        records.append(record)

    result_df = pd.DataFrame(records)
    result_df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    # 標題字體加粗
    title_font = Font(name="Arial", size=14, bold=True)
    for cell in ws[1]:
        cell.font = title_font

    data_font = Font(name="Arial", size=12)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font

    # 設定底色
    color_map = {
        "AF>=0.01": None,
        "AF>=0.4": PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),  # 淡黃
        "AF>=0.5": PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),  # 淡綠
        "AF>=0.6": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # 淡藍
    }

    # 百分比格式 & 底色
    for col_idx, col_name in enumerate(result_df.columns, start=1):
        if "_ratio_" in col_name:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = '0.0%'

            for af_label, fill in color_map.items():
                if af_label in col_name and fill:
                    col_letter = get_column_letter(col_idx)
                    for row_idx in range(2, ws.max_row + 1):
                        ws[f"{col_letter}{row_idx}"].fill = fill

    # 自動調整列寬
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    wb.save(output_file)
    print(f"\n✅ Results saved to {output_file} with ratio %, background color by AF, and auto column width.")

# 🧪 使用範例（請依需求修改路徑）
count_variants(
    '/Users/jasontingting/Desktop/K3 部分結果/',
    '/Users/jasontingting/Desktop/variant_summary_supF(E1)(5).xlsx'
)