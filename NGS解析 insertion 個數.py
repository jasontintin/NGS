import glob
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

def count_variants(folder_path, output_file="variant_summary_Deletion.xlsx"):
    file_list = glob.glob(os.path.join(folder_path, "**/variants_*.tsv"), recursive=True)

    records = []
    total_mutations = 0
    total_insertion = 0
    total_high_af_insertion = 0

    for file in file_list:
        try:
            df = pd.read_csv(file, sep='\t', encoding='utf-8')
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(file, sep='\t', encoding='ISO-8859-1')
            except Exception as e:
                print(f"Error reading {file}: {e}")
                continue

        if df.empty or '34. VARTYPE - variant type' not in df.columns or '15. AF - allele frequency' not in df.columns:
            print(f"Skipping invalid file (missing column): {file}")
            continue

        sample_name = os.path.basename(file).replace("variants_", "").replace(".tsv", "")
        num_variants = len(df)
        num_insertions = (df['34. VARTYPE - variant type'].str.lower() == 'deletion').sum()

        # 新增：篩選符合條件的行
        high_af_insertion_df = df[
            (df['34. VARTYPE - variant type'].str.lower() == 'insertion') &
            (df['15. AF - allele frequency'] > 0.4)
        ]
        high_af_insertion_count = len(high_af_insertion_df)
        total_mutations += num_variants
        total_insertion += num_insertions
        total_high_af_insertion += high_af_insertion_count

        print(
            f"Processed: {sample_name} - Total: {num_variants}, Insertion: {num_insertions}, High AF insertion Samples (>0.5): {high_af_insertion_count}"
        )

        records.append({
            "Sample": sample_name,
            "Total Variants": num_variants,
            "Insertion": num_insertions,
            "High AF Insertion (>0.5)": high_af_insertion_count
        })

    result_df = pd.DataFrame(records)
    result_df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    title_font = Font(name="Arial", size=14, bold=True)
    for cell in ws[1]:
        cell.font = title_font

    data_font = Font(name="Arial", size=12)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font

    wb.save(output_file)
    print(f"\nResults saved to {output_file} with formatted font.")

# 執行
count_variants('/Users/jasontingting/Desktop/K3 部分結果', '/Users/jasontingting/Desktop/variant_summary_insertions_0.4(1).xlsx')