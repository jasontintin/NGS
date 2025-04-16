import glob
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

def count_variants(folder_path, output_file="variant_summary_Deletion.xlsx"):
    file_list = glob.glob(os.path.join(folder_path, "**/variants_*.tsv"), recursive=True)

    records = []

    for file in file_list:
        try:
            df = pd.read_csv(file, sep='\t', encoding='utf-8')
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(file, sep='\t', encoding='ISO-8859-1')
            except Exception as e:
                print(f"Error reading {file}: {e}")
                continue

        if df.empty or '4. Start - start position of the variation' not in df.columns \
            or '34. VARTYPE - variant type' not in df.columns \
            or '15. AF - allele frequency' not in df.columns:
            print(f"Skipping invalid file (missing column): {file}")
            continue

        sample_name = os.path.basename(file).replace("variants_", "").replace(".tsv", "")
        num_variants = len(df)
        num_deletion = (df['34. VARTYPE - variant type'].str.lower() == 'deletion').sum()

        # 條件一：Start >= 97 的數量
        start_ge_97_count = (df['4. Start - start position of the variation'] >= 97).sum()

        # 條件二：Start >= 97 且 AF > 0.4 且 deletion
        filtered_df_af = df[
            (df['4. Start - start position of the variation'] >= 97) &
            (df['15. AF - allele frequency'] > 0.4) &
            (df['34. VARTYPE - variant type'].str.lower() == 'deletion')
        ]
        high_af_deletions_counts = len(filtered_df_af)

        # 條件三：Start >= 97 且 deletion（不看 AF）
        filtered_df_del = df[
            (df['4. Start - start position of the variation'] >= 97) &
            (df['34. VARTYPE - variant type'].str.lower() == 'deletion')
        ]
        start_ge97_deletion_count = len(filtered_df_del)

        # ✅ 條件四：Start >= 97 且 AF >= 0.4（不限 variant type）
        start_af_filtered = df[
            (df['4. Start - start position of the variation'] >= 97) &
            (df['15. AF - allele frequency'] >= 0.4)
        ]
        start_ge97_af_ge04_count = len(start_af_filtered)

        print(
            f"Processed: {sample_name} - Total: {num_variants}, Deletions: {num_deletion}, "
            f"Start≥97: {start_ge_97_count}, Start≥97 & deletion: {start_ge97_deletion_count}, "
            f"Start≥97 & AF>0.4 & deletion: {high_af_deletions_counts}, "
            f"Start≥97 & AF≥0.4: {start_ge97_af_ge04_count}"
        )

        records.append({
            "Sample": sample_name,
            "Total Variants": num_variants,
            "Deletions": num_deletion,
            "Start ≥97 Count": start_ge_97_count,
            "Start ≥97 & Deletion": start_ge97_deletion_count,
            "High AF Deletions (Start ≥97 & AF > 0.4)": high_af_deletions_counts,
            "Start ≥97 & AF ≥0.4": start_ge97_af_ge04_count
        })

    # 輸出到 Excel
    result_df = pd.DataFrame(records)
    result_df.to_excel(output_file, index=False)

    # 套用格式
    wb = load_workbook(output_file)
    ws = wb.active

    title_font = Font(name="Arial", size=14, bold=True)
    for cell in ws[1]:
        cell.font = title_font

    data_font = Font(name="Arial", size=12)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font

    wb.save(output_file)
    print(f"\nResults saved to {output_file} with formatted font.")

# 執行
count_variants(
    '/Users/jasontingting/Desktop/K3 部分結果',
    '/Users/jasontingting/Desktop/variant_summary_Deletiontit.xlsx'
)